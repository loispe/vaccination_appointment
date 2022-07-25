#Author: Louis Petrick 
#Description: 
#Automated script to check for available vaccination appointments.
#
#This basic implementation uses 'selenium' to open and interact with Google Chrome via its 'Chromedriver'.
#By including the desired urls from "impfterminservice.de" to urls.txt, this script will press "Nein (Anspruch Prüfen)" every minute for every included webpage.
#After waiting 5 seconds for the results, the page content will be compared with the one we expect when no appointments are available. If it differs,
#the script will halt and open a pop-up box with the related appointment location.


from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import time
import string
import ctypes


print("***Impftermin Bot V0.0.1***")

def getTime():
    now = datetime.now()
    time_str = now.strftime("%H:%M:%S")
    return [now, time_str]

urls = [
    ["https://229-iz.impfterminservice.de/impftermine/service?plz=89584", "Ehingen", "B"],
    ["https://001-iz.impfterminservice.de/impftermine/service?plz=89073", "Ulm", "C"],
    ["https://003-iz.impfterminservice.de/impftermine/service?plz=89522", "Heidenheim", "D"],
    ["https://003-iz.impfterminservice.de/impftermine/service?plz=88444", "Ummendorf", "E"]
]

wb = load_workbook('log.xlsx')              #excel workbook
ws = wb.create_sheet(str(date.today()))     #excel worksheet

ws['A1'].value = "Uhrzeit"
ws[urls[0][2] + '1'].value = urls[0][1]     #set headings of columns 
ws[urls[1][2] + '1'].value = urls[1][1]
ws[urls[2][2] + '1'].value = urls[2][1]
ws[urls[3][2] + '1'].value = urls[3][1]

print(">Öffne Browser")

browsers = []

for i in range(len(urls)):                  #open one browser for each given url
    browsers.append([webdriver.Chrome(service_log_path='NUL'), urls[i]])
    browsers[i][0].get(urls[i][0])
    time.sleep(1)

wb.save("log.xlsx") 
time.sleep(10)

print(">Suche Termine")

i = 1

for k in range(60*24):
    print("Durchgang #" + str(k))
    i += 1
    warteFlag = False
    now = getTime()
    ws['A' + str(i)].value = now[1]
    
    for browser in browsers:
        try:
            vermittlung_no = browser[0].find_element_by_css_selector("body > app-root > div > app-page-its-login > div > div > div:nth-child(2) > app-its-login-user > div > div > app-corona-vaccination > div:nth-child(2) > div > div > label:nth-child(2)")
            vermittlung_no.click()
            time.sleep(5)
        except:
            warteFlag = True
            time.sleep(1)

        if "Es wurden keine freien Termine in Ihrer Region gefunden" in browser[0].page_source:
            print(now[1] + "\tKein Impfstoff in " + browser[1][1] + "!")
            ws[browser[1][2] + str(i)].value = "-"
        elif warteFlag == True:
            print(now[1] + "\tWarteschlange in " + browser[1][1] + "!")
            ws[browser[1][2] + str(i)].value = "W"
        else:
            print(now[1] + "\tImpfstoff ist Verfügbart in " + browser[1][1] + "!")
            ws[browser[1][2] + str(i)].value = "X" 
            ctypes.windll.user32.MessageBoxW(0, now[1] + "\tImpfstoff ist Verfügbart in " + browser[1][1] + "!", "Impftermin-Bot V0.1", 1)

    print("------------------------------------------------")

    wb.save("log.xlsx")

    while(abs(getTime()[0].minute - now[0].minute) < 1): pass











